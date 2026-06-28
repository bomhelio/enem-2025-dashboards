"""
04_relatorio.py
Gera o relatório Excel consolidado com todos os resultados.
Depende de: output/analise_quantitativa.json + output/benchmark_*.csv
"""

import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR, ESCOLAS, CORTES_NOTA

AREAS = ["CN", "CH", "LC", "MT"]
AREA_NOME = {"CN": "Ciências da Natureza", "CH": "Ciências Humanas",
             "LC": "Linguagens e Códigos", "MT": "Matemática"}
COR_CABECALHO = "1F3864"   # azul escuro
COR_SUBHEADER = "2F75B6"   # azul médio
COR_ALTERNADO = "D9E1F2"   # azul claro para linhas alternadas
COR_DESTAQUE  = "FFEB9C"   # amarelo para linha de benchmarks


# ─── helpers de estilo ────────────────────────────────────────────────────────

def _cabecalho(ws, row, col, texto, cor=COR_CABECALHO, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=texto)
    c.font = Font(bold=bold, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c

def _valor(ws, row, col, valor, bold=False, cor_fundo=None, fmt=None):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if cor_fundo:
        c.fill = PatternFill("solid", fgColor=cor_fundo)
    if fmt:
        c.number_format = fmt
    return c

def _autofit(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(max_len + 2, max_w))


# ─── aba Resumo Executivo ─────────────────────────────────────────────────────

def _aba_resumo(wb: Workbook, dados: dict, benchmarks_geral: pd.DataFrame, benchmarks_priv: pd.DataFrame):
    ws = wb.create_sheet("Resumo Executivo")
    ws.freeze_panes = "B4"

    # Cabeçalho
    ws.merge_cells("A1:O1")
    t = ws["A1"]
    t.value = "ENEM 2025 — Resultado Multimarcas — Resumo Executivo"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    colunas = [
        "Marca", "Unidades", "Estados", "Inscritos", "Presentes 2 dias", "% Presença",
        "Média CN", "Média CH", "Média LC", "Média MT", "Média Redação",
        "Nota Geral", "% ≥600 CN", "% ≥600 CH", "% ≥600 MT",
    ]
    for j, h in enumerate(colunas, 1):
        _cabecalho(ws, 2, j, h, cor=COR_SUBHEADER, wrap=True)
    ws.row_dimensions[2].height = 32

    brasil = benchmarks_geral[benchmarks_geral["SG_UF_ESC"] == "BRASIL"]
    brasil_media = {a: float(brasil[f"media_{a}"].iloc[0]) for a in AREAS}
    brasil_media["REDACAO"] = float(brasil["media_REDACAO"].iloc[0])

    marcas = list(dados.keys())
    for i, marca in enumerate(marcas):
        d = dados[marca]
        row = i + 3
        cor = COR_ALTERNADO if i % 2 == 0 else None
        areas_d = d.get("areas", {})
        red_d   = d.get("redacao", {})
        geral_d = d.get("nota_geral", {})
        valores = [
            marca,
            d.get("unidades", ""),
            ", ".join(d.get("estados", [])),
            d.get("n_inscritos", ""),
            d.get("n_presentes_2dias", ""),
            d.get("taxa_presenca_pct", ""),
            areas_d.get("CN", {}).get("media", ""),
            areas_d.get("CH", {}).get("media", ""),
            areas_d.get("LC", {}).get("media", ""),
            areas_d.get("MT", {}).get("media", ""),
            red_d.get("media", ""),
            geral_d.get("media", ""),
            areas_d.get("CN", {}).get("pct_acima_600", ""),
            areas_d.get("CH", {}).get("pct_acima_600", ""),
            areas_d.get("MT", {}).get("pct_acima_600", ""),
        ]
        for j, v in enumerate(valores, 1):
            _valor(ws, row, j, v, cor_fundo=cor)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # Linha benchmarks Brasil
    row_br = len(marcas) + 3
    _valor(ws, row_br, 1, "BRASIL (todas as redes)", bold=True, cor_fundo=COR_DESTAQUE)
    for j_off, a in enumerate(AREAS, 7):
        _valor(ws, row_br, j_off, brasil_media[a], cor_fundo=COR_DESTAQUE)
    _valor(ws, row_br, 11, brasil_media["REDACAO"], cor_fundo=COR_DESTAQUE)

    _autofit(ws)


# ─── aba por Marca ────────────────────────────────────────────────────────────

def _aba_marca(wb: Workbook, marca: str, d: dict, bench_geral: pd.DataFrame, bench_priv: pd.DataFrame):
    ws = wb.create_sheet(marca[:31])
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"ENEM 2025 — {marca}"
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    row = 2
    areas_d = d.get("areas", {})
    red_d   = d.get("redacao", {})
    geral_d = d.get("nota_geral", {})
    estados = d.get("estados", [])
    bench_estados = bench_geral[bench_geral["SG_UF_ESC"].isin(estados + ["BRASIL"])]

    # Participação
    _cabecalho(ws, row, 1, "PARTICIPAÇÃO", cor=COR_SUBHEADER)
    row += 1
    for label, val in [
        ("Total inscrito", d.get("n_inscritos")),
        ("Presentes nos 2 dias", d.get("n_presentes_2dias")),
        ("Taxa de presença", f"{d.get('taxa_presenca_pct')}%"),
        ("Unidades", d.get("unidades")),
        ("Estados", ", ".join(estados)),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=val).alignment = Alignment(horizontal="left")
        row += 1

    row += 1

    # Notas por área
    _cabecalho(ws, row, 1, "NOTAS POR ÁREA", cor=COR_SUBHEADER)
    row += 1
    hdrs = ["Área", "N", "Média", "Mediana", "DP", "P25", "P75", "P90",
            "Min", "Max"] + [f"% ≥{c}" for c in CORTES_NOTA]
    for j, h in enumerate(hdrs, 1):
        _cabecalho(ws, row, j, h, cor=COR_SUBHEADER)
    row += 1

    for i, (sigla, nome) in enumerate(AREA_NOME.items()):
        s = areas_d.get(sigla, {})
        cor = COR_ALTERNADO if i % 2 == 0 else None
        valores = [
            nome,
            s.get("n"), s.get("media"), s.get("mediana"), s.get("dp"),
            s.get("p25"), s.get("p75"), s.get("p90"), s.get("min"), s.get("max"),
        ] + [s.get(f"pct_acima_{c}") for c in CORTES_NOTA]
        for j, v in enumerate(valores, 1):
            _valor(ws, row, j, v, cor_fundo=cor)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        row += 1

    # Benchmarks para as áreas
    row += 1
    _cabecalho(ws, row, 1, "BENCHMARK (média nacional / privada)", cor=COR_SUBHEADER)
    row += 1
    hdrs_bk = ["Escopo", "N médio"] + [f"Média {a}" for a in AREAS] + ["Média Redação"]
    for j, h in enumerate(hdrs_bk, 1):
        _cabecalho(ws, row, j, h, cor=COR_SUBHEADER)
    row += 1

    for df_bench, label in [(bench_geral, "Nacional"), (bench_priv, "Privada Nacional")]:
        br = df_bench[df_bench["SG_UF_ESC"] == "BRASIL"]
        if br.empty:
            continue
        br = br.iloc[0]
        vals = [label, ""] + [float(br[f"media_{a}"]) for a in AREAS] + [float(br["media_REDACAO"])]
        for j, v in enumerate(vals, 1):
            _valor(ws, row, j, v, cor_fundo=COR_DESTAQUE)
        row += 1

    row += 1

    # Redação
    _cabecalho(ws, row, 1, "REDAÇÃO", cor=COR_SUBHEADER)
    row += 1
    ws.cell(row=row, column=1, value="Redações válidas").font = Font(bold=True, size=10)
    ws.cell(row=row, column=2, value=red_d.get("n_validas"))
    row += 1
    ws.cell(row=row, column=1, value="% Zero / Anuladas").font = Font(bold=True, size=10)
    ws.cell(row=row, column=2, value=red_d.get("pct_nota_zero_anulada"))
    row += 1
    ws.cell(row=row, column=1, value="Média total").font = Font(bold=True, size=10)
    ws.cell(row=row, column=2, value=red_d.get("media"))
    row += 2

    # Competências
    _cabecalho(ws, row, 1, "Competências da Redação", cor=COR_SUBHEADER)
    row += 1
    for j, h in enumerate(["Competência", "Descrição", "N", "Média", "Mediana", "DP", "P25", "P75"], 1):
        _cabecalho(ws, row, j, h, cor=COR_SUBHEADER)
    row += 1
    for num in range(1, 6):
        s = red_d.get("competencias", {}).get(f"comp{num}", {})
        cor = COR_ALTERNADO if num % 2 == 0 else None
        vals = [f"Comp {num}", s.get("nome", ""), s.get("n"), s.get("media"),
                s.get("mediana"), s.get("dp"), s.get("p25"), s.get("p75")]
        for j, v in enumerate(vals, 1):
            _valor(ws, row, j, v, cor_fundo=cor)
        row += 1

    row += 1

    # Distribuição de faixas da redação
    _cabecalho(ws, row, 1, "Distribuição por Faixa — Redação", cor=COR_SUBHEADER)
    row += 1
    dist = red_d.get("distribuicao_faixas", {})
    n_validas = red_d.get("n_validas", 0) or 1
    for j, h in enumerate(["Faixa", "Quantidade", "% do Total Válido"], 1):
        _cabecalho(ws, row, j, h, cor=COR_SUBHEADER)
    row += 1
    for i, (faixa, cnt) in enumerate(dist.items()):
        cor = COR_ALTERNADO if i % 2 == 0 else None
        _valor(ws, row, 1, faixa, cor_fundo=cor)
        _valor(ws, row, 2, cnt, cor_fundo=cor)
        _valor(ws, row, 3, round(cnt / n_validas * 100, 1), cor_fundo=cor)
        row += 1

    row += 1

    # Nota Geral
    _cabecalho(ws, row, 1, "NOTA GERAL (média das 5 áreas — apenas presença completa + redação válida)", cor=COR_SUBHEADER)
    row += 1
    for label, key in [("N", "n_completos"), ("Média", "media"), ("Mediana", "mediana"),
                        ("DP", "dp"), ("P25", "p25"), ("P75", "p75"), ("P90", "p90")]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=geral_d.get(key))
        row += 1

    _autofit(ws)


# ─── aba Comparativo ──────────────────────────────────────────────────────────

def _aba_comparativo(wb: Workbook, dados: dict, benchmarks_geral: pd.DataFrame):
    ws = wb.create_sheet("Comparativo Marcas")
    ws.freeze_panes = "B3"

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Comparativo entre Marcas — ENEM 2025"
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COR_CABECALHO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    metricas = [
        ("Média CN",     lambda d: d.get("areas", {}).get("CN", {}).get("media")),
        ("Média CH",     lambda d: d.get("areas", {}).get("CH", {}).get("media")),
        ("Média LC",     lambda d: d.get("areas", {}).get("LC", {}).get("media")),
        ("Média MT",     lambda d: d.get("areas", {}).get("MT", {}).get("media")),
        ("Média Redação",lambda d: d.get("redacao", {}).get("media")),
        ("Nota Geral",   lambda d: d.get("nota_geral", {}).get("media")),
        ("% Presença",   lambda d: d.get("taxa_presenca_pct")),
        ("% Zer/Anul Redação", lambda d: d.get("redacao", {}).get("pct_nota_zero_anulada")),
    ] + [
        (f"% ≥{c} CN", lambda d, c=c: d.get("areas", {}).get("CN", {}).get(f"pct_acima_{c}"))
        for c in CORTES_NOTA
    ] + [
        (f"% ≥{c} MT", lambda d, c=c: d.get("areas", {}).get("MT", {}).get(f"pct_acima_{c}"))
        for c in CORTES_NOTA
    ]

    marcas = list(dados.keys())
    _cabecalho(ws, 2, 1, "Métrica", cor=COR_SUBHEADER)
    for j, m in enumerate(marcas, 2):
        _cabecalho(ws, 2, j, m, cor=COR_SUBHEADER)
    _cabecalho(ws, 2, len(marcas) + 2, "BRASIL", cor=COR_SUBHEADER)

    brasil = benchmarks_geral[benchmarks_geral["SG_UF_ESC"] == "BRASIL"]
    benchmarks_brasil = {
        "Média CN": float(brasil["media_CN"].iloc[0]) if not brasil.empty else None,
        "Média CH": float(brasil["media_CH"].iloc[0]) if not brasil.empty else None,
        "Média LC": float(brasil["media_LC"].iloc[0]) if not brasil.empty else None,
        "Média MT": float(brasil["media_MT"].iloc[0]) if not brasil.empty else None,
        "Média Redação": float(brasil["media_REDACAO"].iloc[0]) if not brasil.empty else None,
    }

    for i, (nome_metrica, fn) in enumerate(metricas):
        row = i + 3
        cor = COR_ALTERNADO if i % 2 == 0 else None
        _valor(ws, row, 1, nome_metrica, bold=True, cor_fundo=cor)
        for j, marca in enumerate(marcas, 2):
            _valor(ws, row, j, fn(dados[marca]), cor_fundo=cor)
        _valor(ws, row, len(marcas) + 2, benchmarks_brasil.get(nome_metrica), cor_fundo=COR_DESTAQUE)

    _autofit(ws)


# ─── main ─────────────────────────────────────────────────────────────────────

def gerar_relatorio():
    analise_path = os.path.join(OUTPUT_DIR, "analise_quantitativa.json")
    bench_geral_path = os.path.join(OUTPUT_DIR, "benchmark_geral.csv")
    bench_priv_path  = os.path.join(OUTPUT_DIR, "benchmark_privada.csv")

    if not os.path.exists(analise_path):
        raise FileNotFoundError(f"Rode 02_quantitativo.py antes. Não encontrado: {analise_path}")

    with open(analise_path, encoding="utf-8") as f:
        dados = json.load(f)

    bench_geral = pd.read_csv(bench_geral_path) if os.path.exists(bench_geral_path) else pd.DataFrame()
    bench_priv  = pd.read_csv(bench_priv_path)  if os.path.exists(bench_priv_path)  else pd.DataFrame()

    wb = Workbook()
    wb.remove(wb.active)  # remove sheet padrão

    print("  Gerando aba Resumo Executivo...")
    _aba_resumo(wb, dados, bench_geral, bench_priv)

    print("  Gerando aba Comparativo...")
    _aba_comparativo(wb, dados, bench_geral)

    for marca in dados:
        print(f"  Gerando aba {marca}...")
        _aba_marca(wb, marca, dados[marca], bench_geral, bench_priv)

    destino = os.path.join(OUTPUT_DIR, "ENEM_2025_Multimarcas.xlsx")
    wb.save(destino)
    print(f"\nRelatório salvo: {destino}")
    return destino


if __name__ == "__main__":
    gerar_relatorio()
